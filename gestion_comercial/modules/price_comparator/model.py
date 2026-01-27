"""
Modelo para comparación de precios entre bases de datos
"""

import openpyxl
import os
from datetime import datetime
from typing import Dict, List, Tuple, Optional


class PriceComparator:
    """Maneja la comparación de precios entre dos bases de datos Excel"""

    def __init__(self):
        self.main_db_path = None
        self.comparison_db_path = None
        self.main_db_date = None
        self.comparison_db_date = None
        self.main_data = {}  # {barcode: {'name': str, 'price': float, 'department': str}}
        self.comparison_data = {}
        self.differences = []  # Lista de diferencias encontradas

    def load_database(self, file_path: str) -> Dict[str, dict]:
        """
        Carga una base de datos Excel y retorna un diccionario
        {codigo_barras: {'name': nombre, 'price': precio, 'department': departamento}}
        """
        data = {}
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            ws = wb.active

            # Comenzar desde la fila 2 si la primera fila tiene encabezados
            # O desde la fila 1 si no hay encabezados
            start_row = 1

            # Verificar si la primera fila contiene encabezados
            first_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
            print(f"[DEBUG] Primera fila del archivo: {first_row[:5] if len(first_row) >= 5 else first_row}")

            if first_row and len(first_row) >= 4:
                # Si el cuarto valor (P. Venta) no es numérico, probablemente es un encabezado
                try:
                    test_value = first_row[3]
                    print(f"[DEBUG] Valor en columna D (P.Venta): '{test_value}' (tipo: {type(test_value)})")

                    # Intentar convertir, manejando strings que empiecen con $
                    if isinstance(test_value, str):
                        test_clean = test_value.replace('$', '').replace(',', '').strip()
                        float(test_clean)
                    else:
                        float(test_value)
                    print(f"[DEBUG] Es numérico, NO saltar encabezados")
                except (ValueError, TypeError) as e:
                    start_row = 2  # Saltar la fila de encabezados
                    print(f"[DEBUG] NO es numérico ({e}), saltar fila de encabezados")

            print(f"[DEBUG] Comenzando lectura desde fila {start_row}")
            row_count = 0
            for row in ws.iter_rows(min_row=start_row, values_only=True):
                row_count += 1
                if row_count <= 3:  # Mostrar primeras 3 filas para debug
                    print(f"[DEBUG] Fila {start_row + row_count - 1}: {row[:7] if len(row) >= 7 else row}")

                # Verificar que tengamos al menos 4 columnas (Código, Producto, P.Costo, P.Venta)
                if len(row) >= 4 and row[0] and row[1] and row[3]:
                    barcode = str(row[0]).strip()
                    name = str(row[1]).strip()

                    # Leer departamento (columna F, índice 5) si existe
                    department = ''
                    if len(row) >= 6 and row[5]:
                        department = str(row[5]).strip()

                    try:
                        # Leer P. Venta (columna D, índice 3)
                        price_value = row[3]

                        # Si viene como string con formato ($X.XXX), limpiar
                        if isinstance(price_value, str):
                            price_clean = price_value.replace('$', '').replace(',', '').replace('.', '').strip()
                            price = float(price_clean)
                        else:
                            price = float(price_value)

                    except (ValueError, TypeError) as e:
                        if row_count <= 3:
                            print(f"[DEBUG] Error al procesar precio en fila {start_row + row_count - 1}: {e}")
                        continue

                    data[barcode] = {
                        'name': name,
                        'price': price,
                        'department': department
                    }

            print(f"[DEBUG] Total filas procesadas: {row_count}, Productos cargados: {len(data)}")

            wb.close()
            return data

        except Exception as e:
            raise Exception(f"Error al cargar la base de datos: {str(e)}")

    def set_main_database(self, file_path: str):
        """Establece y carga la base de datos principal"""
        self.main_db_path = file_path
        self.main_db_date = datetime.fromtimestamp(os.path.getmtime(file_path))
        self.main_data = self.load_database(file_path)
        print(f"[DEBUG] Base principal cargada: {len(self.main_data)} productos")
        if len(self.main_data) > 0:
            # Mostrar primeros 3 productos
            for i, (barcode, info) in enumerate(list(self.main_data.items())[:3]):
                print(f"  [{i+1}] {barcode}: {info['name']} - ${info['price']:,.0f}")

    def set_comparison_database(self, file_path: str):
        """Establece y carga la base de datos de comparación"""
        self.comparison_db_path = file_path
        self.comparison_db_date = datetime.fromtimestamp(os.path.getmtime(file_path))
        self.comparison_data = self.load_database(file_path)
        print(f"[DEBUG] Base de comparación cargada: {len(self.comparison_data)} productos")
        if len(self.comparison_data) > 0:
            # Mostrar primeros 3 productos
            for i, (barcode, info) in enumerate(list(self.comparison_data.items())[:3]):
                print(f"  [{i+1}] {barcode}: {info['name']} - ${info['price']:,.0f}")

    def compare_databases(self) -> List[dict]:
        """
        Compara las dos bases de datos y retorna una lista de diferencias

        Retorna lista de diccionarios con:
        {
            'barcode': str,
            'name': str,
            'main_price': float,
            'comparison_price': float or None,
            'difference': float,
            'status': 'price_diff' | 'missing'
        }
        """
        differences = []

        print(f"\n[DEBUG] Iniciando comparación...")
        print(f"  Productos en base principal: {len(self.main_data)}")
        print(f"  Productos en base comparación: {len(self.comparison_data)}")

        # Recorrer todos los productos de la base principal
        for barcode, main_info in self.main_data.items():

            if barcode in self.comparison_data:
                # Producto existe en ambas bases
                comp_info = self.comparison_data[barcode]

                # Comparar precios
                if main_info['price'] != comp_info['price']:
                    diff = main_info['price'] - comp_info['price']
                    print(f"  [DIFF] {barcode}: ${main_info['price']:,.0f} vs ${comp_info['price']:,.0f} (diferencia: ${diff:,.0f})")
                    differences.append({
                        'barcode': barcode,
                        'name': main_info['name'],
                        'main_price': main_info['price'],
                        'comparison_price': comp_info['price'],
                        'difference': diff,
                        'status': 'price_diff'
                    })
            else:
                # Producto falta en la base de comparación
                print(f"  [MISSING] {barcode}: {main_info['name']} (falta en comparación)")
                differences.append({
                    'barcode': barcode,
                    'name': main_info['name'],
                    'main_price': main_info['price'],
                    'comparison_price': None,
                    'difference': None,
                    'status': 'missing'
                })

        print(f"\n[DEBUG] Comparación completada:")
        print(f"  Total diferencias encontradas: {len(differences)}")
        self.differences = differences
        return differences

    def update_price(self, barcode: str, new_price: float) -> bool:
        """
        Actualiza el precio de un producto en la base de comparación
        """
        try:
            wb = openpyxl.load_workbook(self.comparison_db_path)
            ws = wb.active

            # Buscar la fila con el código de barras
            for row_idx, row in enumerate(ws.iter_rows(min_row=1, values_only=False), start=1):
                if str(row[0].value).strip() == barcode:
                    # Actualizar el precio (columna D - P. Venta)
                    ws.cell(row=row_idx, column=4, value=new_price)
                    wb.save(self.comparison_db_path)
                    wb.close()

                    # Actualizar datos en memoria
                    if barcode in self.comparison_data:
                        self.comparison_data[barcode]['price'] = new_price

                    return True

            wb.close()
            return False

        except Exception as e:
            raise Exception(f"Error al actualizar precio: {str(e)}")

    def add_product(self, barcode: str) -> bool:
        """
        Agrega un producto faltante a la base de comparación
        """
        try:
            if barcode not in self.main_data:
                return False

            main_info = self.main_data[barcode]

            wb = openpyxl.load_workbook(self.comparison_db_path)
            ws = wb.active

            # Agregar nueva fila al final con todas las columnas
            # Columnas: Código, Producto, P.Costo, P.Venta, (resto vacíos)
            new_row = [barcode, main_info['name'], '$0', main_info['price']]
            ws.append(new_row)

            wb.save(self.comparison_db_path)
            wb.close()

            # Actualizar datos en memoria
            self.comparison_data[barcode] = {
                'name': main_info['name'],
                'price': main_info['price']
            }

            return True

        except Exception as e:
            raise Exception(f"Error al agregar producto: {str(e)}")

    def get_statistics(self) -> dict:
        """Retorna estadísticas de la comparación"""
        if not self.differences:
            return {
                'total_differences': 0,
                'price_differences': 0,
                'missing_products': 0
            }

        price_diffs = len([d for d in self.differences if d['status'] == 'price_diff'])
        missing = len([d for d in self.differences if d['status'] == 'missing'])

        return {
            'total_differences': len(self.differences),
            'price_differences': price_diffs,
            'missing_products': missing
        }

    def generate_report(self) -> str:
        """
        Genera un reporte completo en formato de texto con toda la información de la comparación
        """
        lines = []
        lines.append("=" * 80)
        lines.append("REPORTE DE COMPARACIÓN DE PRECIOS")
        lines.append("=" * 80)
        lines.append("")

        # RESUMEN GENERAL
        lines.append("RESUMEN GENERAL")
        lines.append("-" * 80)

        # Nombres de archivos
        main_filename = os.path.basename(self.main_db_path) if self.main_db_path else "N/A"
        comp_filename = os.path.basename(self.comparison_db_path) if self.comparison_db_path else "N/A"

        lines.append(f"Base de datos principal: {main_filename}")
        if self.main_db_date:
            lines.append(f"  Fecha: {self.main_db_date.strftime('%d/%m/%Y %H:%M:%S')}")
        lines.append(f"  Total de productos: {len(self.main_data)}")
        lines.append("")

        lines.append(f"Base de datos de comparación: {comp_filename}")
        if self.comparison_db_date:
            lines.append(f"  Fecha: {self.comparison_db_date.strftime('%d/%m/%Y %H:%M:%S')}")
        lines.append(f"  Total de productos: {len(self.comparison_data)}")
        lines.append("")

        # Estadísticas
        stats = self.get_statistics()
        lines.append(f"Total de diferencias detectadas: {stats['total_differences']}")
        lines.append(f"  - Diferencias de precio: {stats['price_differences']}")
        lines.append(f"  - Productos faltantes: {stats['missing_products']}")
        lines.append("")
        lines.append("")

        # PRODUCTOS FALTANTES
        missing_products = [d for d in self.differences if d['status'] == 'missing']
        if missing_products:
            lines.append("=" * 80)
            lines.append("PRODUCTOS FALTANTES EN BASE DE COMPARACIÓN")
            lines.append("=" * 80)
            lines.append("")
            lines.append(f"{'Código':<15} {'Producto':<35} {'Precio':<12} {'Departamento'}")
            lines.append("-" * 80)

            for prod in missing_products:
                barcode = prod['barcode']
                name = prod['name'][:35]  # Limitar a 35 caracteres
                price = f"${prod['main_price']:,.0f}"
                department = self.main_data[barcode].get('department', 'N/A')

                lines.append(f"{barcode:<15} {name:<35} {price:<12} {department}")

            lines.append("")
            lines.append("")

        # PRODUCTOS CON PRECIOS MODIFICADOS
        price_diffs = [d for d in self.differences if d['status'] == 'price_diff']
        if price_diffs:
            lines.append("=" * 80)
            lines.append("PRODUCTOS CON DIFERENCIAS DE PRECIO")
            lines.append("=" * 80)
            lines.append("")
            lines.append(f"{'Código':<15} {'Producto':<25} {'Precio Actual':<15} {'Precio Comp.':<15} {'Precio Sugerido'}")
            lines.append("-" * 80)

            for prod in price_diffs:
                barcode = prod['barcode']
                name = prod['name'][:25]  # Limitar a 25 caracteres
                main_price = f"${prod['main_price']:,.0f}"
                comp_price = f"${prod['comparison_price']:,.0f}"

                # Calcular precio sugerido (promedio)
                suggested_price = (prod['main_price'] + prod['comparison_price']) / 2
                suggested_price_str = f"${suggested_price:,.0f}"

                lines.append(f"{barcode:<15} {name:<25} {main_price:<15} {comp_price:<15} {suggested_price_str}")

            lines.append("")
            lines.append("")

        # PIE DE PÁGINA
        lines.append("=" * 80)
        lines.append(f"Reporte generado: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")
        lines.append("=" * 80)

        return "\n".join(lines)
