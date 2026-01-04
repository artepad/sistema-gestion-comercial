"""
Módulo de gestión de base de datos de productos.
Permite buscar productos por código de barras desde un archivo Excel.
"""

import os
import glob
from datetime import datetime
from openpyxl import load_workbook


class ProductDatabase:
    """Gestiona la base de datos de productos en formato Excel."""

    # Ruta de la carpeta de base de datos
    DB_FOLDER = os.path.join(
        os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))),
        'bd'
    )

    @staticmethod
    def get_db_file():
        """
        Busca el primer archivo .xlsx en la carpeta bd.

        Returns:
            str: Ruta del archivo encontrado o None
        """
        if not os.path.exists(ProductDatabase.DB_FOLDER):
            return None

        # Buscar archivos .xlsx en la carpeta
        xlsx_files = glob.glob(os.path.join(ProductDatabase.DB_FOLDER, '*.xlsx'))

        if xlsx_files:
            # Retornar el primer archivo encontrado
            return xlsx_files[0]

        return None

    @staticmethod
    def file_exists():
        """
        Verifica si existe algún archivo de base de datos.

        Returns:
            bool: True si existe, False en caso contrario
        """
        return ProductDatabase.get_db_file() is not None

    @staticmethod
    def get_file_date():
        """
        Obtiene la fecha de última modificación del archivo.

        Returns:
            str: Fecha formateada o mensaje de error
        """
        db_file = ProductDatabase.get_db_file()

        if not db_file:
            return "Archivo no encontrado"

        try:
            timestamp = os.path.getmtime(db_file)
            date_obj = datetime.fromtimestamp(timestamp)
            return date_obj.strftime("%d/%m/%Y %H:%M")
        except Exception as e:
            return f"Error al leer fecha: {e}"

    @staticmethod
    def search_product(barcode):
        """
        Busca un producto por código de barras.

        Args:
            barcode (str): Código de barras del producto

        Returns:
            tuple: (success: bool, result: dict or error_message: str)
                   Si success=True, result contiene {'name': str, 'price': float}
                   Si success=False, result contiene mensaje de error
        """
        # Obtener archivo de BD
        db_file = ProductDatabase.get_db_file()

        if not db_file:
            return False, "El archivo de base de datos no existe.\nUbicación esperada: gestion_comercial/bd/*.xlsx"

        # Validar código de barras
        if not barcode or not barcode.strip():
            return False, "Código de barras vacío"

        try:
            # Cargar el archivo Excel
            workbook = load_workbook(db_file, read_only=True, data_only=True)
            sheet = workbook.active

            # Buscar en las filas
            # Estructura del archivo:
            # Columna A (0): Código
            # Columna B (1): Producto
            # Columna D (3): P. Venta
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if not row or len(row) < 4:
                    continue

                # Comparar código de barras (convertir a string para comparación)
                db_code = str(row[0]).strip() if row[0] else ""
                search_code = str(barcode).strip()

                if db_code == search_code:
                    # Producto encontrado
                    product_name = str(row[1]).strip() if row[1] else ""

                    # Obtener precio de venta (columna D, índice 3)
                    price_str = str(row[3]).strip() if row[3] else "0"

                    # Limpiar el precio (quitar $, puntos de miles, espacios)
                    price_str = price_str.replace('$', '').replace('.', '').replace(',', '.').replace(' ', '')

                    # Intentar convertir precio a float
                    try:
                        product_price = float(price_str) if price_str else 0.0
                    except (ValueError, TypeError):
                        product_price = 0.0

                    workbook.close()

                    return True, {
                        'name': product_name,
                        'price': product_price
                    }

            # No se encontró el producto
            workbook.close()
            return False, f"Producto con código '{barcode}' no encontrado en la base de datos"

        except Exception as e:
            return False, f"Error al leer la base de datos: {str(e)}"

    @staticmethod
    def get_database_info():
        """
        Obtiene información general de la base de datos.

        Returns:
            dict: Información de la base de datos
        """
        db_file = ProductDatabase.get_db_file()

        info = {
            'exists': db_file is not None,
            'path': db_file if db_file else os.path.join(ProductDatabase.DB_FOLDER, '*.xlsx'),
            'last_modified': ProductDatabase.get_file_date()
        }

        if info['exists']:
            try:
                workbook = load_workbook(db_file, read_only=True)
                sheet = workbook.active
                # Contar filas (excluyendo encabezado)
                info['total_products'] = sheet.max_row - 1 if sheet.max_row > 1 else 0
                workbook.close()
            except Exception:
                info['total_products'] = 0
        else:
            info['total_products'] = 0

        return info


if __name__ == "__main__":
    # Prueba del módulo
    print("=== Test del modulo de base de datos ===\n")

    info = ProductDatabase.get_database_info()
    print(f"Archivo existe: {info['exists']}")
    print(f"Ruta: {info['path']}")
    print(f"Ultima modificacion: {info['last_modified']}")
    print(f"Total de productos: {info['total_products']}")

    if info['exists']:
        print("\nBuscando producto de prueba (1100000100 - Manzana)...")
        success, result = ProductDatabase.search_product("1100000100")
        if success:
            print(f"OK - Producto encontrado:")
            print(f"  Nombre: {result['name']}")
            print(f"  Precio: ${result['price']}")
        else:
            print(f"ERROR - {result}")
